import requests
import json
import xmltodict
import datetime
import calendar
import os.path
import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook

class Dolar(object):

    def __init__(self):
        self._url = 'https://www3.bcb.gov.br/bc_moeda/rest/converter/1/1/790/220/'

    def get_value(self,year,month,day):
        header = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36'}
        page = requests.get(f'{self._url}{year}-{month}-{day}', headers=header)
        
        if page.status_code == 200:
            valor = json.loads(json.dumps(xmltodict.parse(page.content)))

            print(valor['valor-convertido'])
            return f'{day}-{month}-{year}' , "%.4f" %(1 / float(valor['valor-convertido']))

        return '','--' # dia sem fechamento de cotação

    def get_range_datas(self):
        month = datetime.datetime.today().month -1
        year = datetime.datetime.today().year
        last_day = calendar.monthrange(year, month)[1]

        return year,month,last_day

    def get_dict_data(self, year,month,last_day):
        valores = dict()
        i = 1
        while i <= last_day:
            data,valor = self.get_value(year,month,i)
            if valor != "--":
                valores[data] = valor
                print('.')
            i+=1

        return valores

class Arquivo(object):
    
    def __init__(self):
        self._local = Path('C:/CotacaoDolar')
        self._path_txt = Path('C:/CotacaoDolar/Log.txt')
        self._path_xlsx = Path('C:/CotacaoDolar/Valores.xlsx')

        if not os.path.exists(self._local):
            os.makedirs(self._local)
            print('criou o arquivo')

        self.arquivo_txt = open(self._path_txt, 'w')
        self.arquivo_txt.write('Início da execução \n')

    def set_valores_log(self, dict):
        lower = 9999.99
        higher = -1.99
        sum = 0        
        
        if dict:
            for data,value in dict.items():
                aux_value = float(value)
                if aux_value < lower:
                    lower = aux_value
                
                if aux_value > higher:
                    higher = aux_value  

                sum += aux_value
        
        self.arquivo_txt.write(f'Variação de valores no período:\n')
        self.arquivo_txt.write(f'Máxima - {higher}\n')
        self.arquivo_txt.write(f'Mínima - {lower}\n')
        self.arquivo_txt.write(f'Média - {"%.4f" % float(sum/len(dict))} \n')
        
    def salva_txt(self):
        self.arquivo_txt.write(f'Fim da execução - {datetime.datetime.now()}')
        self.arquivo_txt.close()
        print('Arquivo gerado em C:/CotacaoDolar')

    def build_xlsx(self, dict):
        
        wb = Workbook()
        dest_filename = self._path_xlsx

        ws3 = wb.active 
        ws3.title = "Cotação Dólar"

        index = 0
        for data,valor in dict.items():
            index += 1
            _ = ws3.cell(column=1, row=index, value=data)
            _ = ws3.cell(column=2, row=index, value=valor)
        
        wb.save(filename = dest_filename)

    def load_from_xlsx(self):

        dfs = pd.read_excel(self._path_xlsx, sheet_name=None, header=None)

        novo = dfs['Cotação Dólar']

        dc = {}

        i = 0
        while(i < len(novo[0])):
            dc[str(novo[0][i])] = str(novo[1][i])
            i = i + 1

        return dc